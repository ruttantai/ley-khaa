"""The offline lane end of the synthesizer contract.

These tests run the canned scripts for real, in the real sandbox, against the
real catalog. The heuristic parses the prompt the Synthesizer builds, so a
change to either side that breaks the pair fails here rather than silently
turning the no-API-key demo into an empty bundle.
"""
from openpyxl import load_workbook

from ley_khaa.executor import catalog
from ley_khaa.executor.resolver import ResolvedInput
from ley_khaa.executor.sandbox import SubprocessSandbox
from ley_khaa.executor.synthesizer import Synthesizer
from ley_khaa.executor.workspace import Workspace
from ley_khaa.interpreter.spec import TaskSpec
from ley_khaa.llm.heuristic import HeuristicLLM


def _spec(operation="set_difference", output_format="xlsx") -> TaskSpec:
    return TaskSpec(
        intent="find what is missing",
        inputs=["Bloomberg universe", "FactSet"],
        operation=operation,
        output_format=output_format,
        certainty=0.55,
    )


def _universes() -> list[ResolvedInput]:
    return [
        ResolvedInput(
            name="Bloomberg universe",
            filename="bloomberg_universe.csv",
            content=catalog.build_dataset("bloomberg_universe"),
            source="catalog",
        ),
        ResolvedInput(
            name="FactSet",
            filename="factset_universe.csv",
            content=catalog.build_dataset("factset_universe"),
            source="catalog",
        ),
    ]


def _run(tmp_path, spec, resolved):
    """Synthesize offline, then actually execute the result."""
    script = Synthesizer(HeuristicLLM()).synthesize(spec, resolved)
    workspace = Workspace.create(tmp_path, "t1")
    workspace.write_inputs(resolved)
    path = workspace.write_generator(1, script.source)
    result = SubprocessSandbox().run(script=path, workspace=workspace.root, timeout_s=60)
    return workspace, result


def test_the_offline_set_difference_actually_writes_a_spreadsheet(tmp_path):
    workspace, result = _run(tmp_path, _spec(), _universes())
    assert result.ok, result.stderr
    book = load_workbook(workspace.deliverable_dir / "output.xlsx")
    # Task 1 guarantees bloomberg has exactly 5 tickers factset lacks; header + 5.
    assert book.active.max_row == 6
    assert [cell.value for cell in book.active[1]][0] == "ticker"


def test_the_offline_lane_honours_a_csv_request(tmp_path):
    workspace, result = _run(tmp_path, _spec(output_format="csv"), _universes())
    assert result.ok, result.stderr
    lines = (workspace.deliverable_dir / "output.csv").read_text().splitlines()
    assert len(lines) == 6


def test_summary_stats_describes_the_numeric_columns(tmp_path):
    resolved = [
        ResolvedInput(
            name="holdings",
            filename="holdings.csv",
            content=catalog.build_dataset("holdings"),
            source="catalog",
        )
    ]
    spec = _spec(operation="summary_stats", output_format="csv")
    workspace, result = _run(tmp_path, spec, resolved)
    assert result.ok, result.stderr
    body = (workspace.deliverable_dir / "output.csv").read_text()
    assert body.startswith("column,count,min,max,mean\n")
    assert "quantity" in body


def test_an_unrecognised_operation_still_produces_an_honest_deliverable(tmp_path):
    """The offline stand-in must never leave the fresh-clone demo with no
    bundle. A request it cannot pattern-match gets a truthful description of the
    inputs rather than an exception."""
    spec = _spec(operation="reconcile_everything", output_format="csv")
    workspace, result = _run(tmp_path, spec, _universes())
    assert result.ok, result.stderr
    body = (workspace.deliverable_dir / "output.csv").read_text()
    assert "bloomberg_universe.csv" in body


def test_the_canned_script_never_touches_its_inputs(tmp_path):
    """Reproducibility is the bundle's whole claim, and the validator enforces
    it — the offline lane must not be the thing that trips it."""
    workspace, result = _run(tmp_path, _spec(), _universes())
    assert result.ok, result.stderr
    before = {
        item.filename: item.sha256 for item in _universes()
    }
    assert workspace.input_hashes() == before


def test_the_offline_script_is_deterministic(tmp_path):
    first, _ = _run(tmp_path / "a", _spec(output_format="csv"), _universes())
    second, _ = _run(tmp_path / "b", _spec(output_format="csv"), _universes())
    assert (first.deliverable_dir / "output.csv").read_bytes() == (
        second.deliverable_dir / "output.csv"
    ).read_bytes()
