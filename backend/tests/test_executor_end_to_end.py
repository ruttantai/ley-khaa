"""§9: a fresh clone with no ANTHROPIC_API_KEY produces a real spreadsheet.

Nothing here is mocked. The heuristic LLM synthesizes, the subprocess sandbox
runs it, the validator judges it, and the bundle lands on disk.
"""
import csv
import json
from pathlib import Path

from openpyxl import load_workbook

from ley_khaa.executor.sandbox import SubprocessSandbox


def _cells(path: Path) -> list[tuple]:
    book = load_workbook(path)
    try:
        return [tuple(row) for row in book.active.iter_rows(values_only=True)]
    finally:
        book.close()


def _run_the_golden_conversation(client) -> dict:
    client.post("/simulate/messy_universe_check")
    task = client.get("/tasks").json()[0]
    return client.post(f"/tasks/{task['id']}/mode", json={"mode": "auto"}).json()


def test_the_golden_conversation_produces_a_real_spreadsheet(client):
    task = _run_the_golden_conversation(client)
    assert task["state"] == "done"

    root = Path(task["workspace_path"])
    deliverable = root / "deliverable" / "output.xlsx"
    assert deliverable.is_file()
    # Task 1 guarantees bloomberg holds exactly 5 tickers factset lacks.
    assert len(_cells(deliverable)) == 6


def test_the_bundle_records_how_the_spreadsheet_was_made(client):
    task = _run_the_golden_conversation(client)
    root = Path(task["workspace_path"])
    manifest = json.loads((root / "manifest.json").read_text())

    assert manifest["lane"] == "synthesis"
    # conftest pins the subprocess sandbox; the manifest must say so rather than
    # claiming the isolation it did not have.
    assert manifest["sandbox"] == "subprocess"
    assert manifest["verdict"]["ok"] is True
    assert [i["file"] for i in manifest["inputs"]] == [
        "bloomberg_universe.csv",
        "factset_universe.csv",
    ]
    assert len(manifest["deliverables"][0]["sha256"]) == 64
    assert (root / "generator" / "attempt_1.py").is_file()
    assert (root / "generator" / "run.sh").is_file()


def test_the_bundle_re_runs_to_the_same_spreadsheet(client):
    """The claim the whole Output Bundle rests on."""
    task = _run_the_golden_conversation(client)
    root = Path(task["workspace_path"])
    deliverable = root / "deliverable" / "output.xlsx"

    original = _cells(deliverable)
    deliverable.unlink()

    script = root / "generator" / "attempt_1.py"
    result = SubprocessSandbox().run(script=script, workspace=root, timeout_s=60)
    assert result.ok, result.stderr
    # Values, not bytes: an .xlsx is a zip and embeds a timestamp.
    assert _cells(deliverable) == original


def test_a_csv_bundle_re_runs_byte_for_byte(client):
    """Where a byte-level claim IS available, make it."""
    client.post("/simulate/ambiguous_report_request")
    task = next(
        t for t in client.get("/tasks").json() if t["state"] == "needs_clarification"
    )
    client.post(f"/tasks/{task['id']}/answer", json={"text": "as a csv please"})
    done = client.post(f"/tasks/{task['id']}/approve").json()
    assert done["state"] == "done"

    root = Path(done["workspace_path"])
    deliverable = root / "deliverable" / "output.csv"
    original = deliverable.read_bytes()
    assert len(list(csv.reader(original.decode().splitlines()))) > 1

    deliverable.unlink()
    result = SubprocessSandbox().run(
        script=root / "generator" / "attempt_1.py", workspace=root, timeout_s=60
    )
    assert result.ok, result.stderr
    assert deliverable.read_bytes() == original


def test_the_dashboard_can_reach_the_bundle_over_the_api(client):
    task = _run_the_golden_conversation(client)
    bundle = client.get(f"/tasks/{task['id']}/bundle").json()
    assert bundle["deliverables"] == ["deliverable/output.xlsx"]

    source = client.get(
        f"/tasks/{task['id']}/bundle/file", params={"path": "generator/attempt_1.py"}
    ).json()["content"]
    assert "write_rows" in source

    assert client.get(f"/tasks/{task['id']}/bundle/deliverable").status_code == 200
