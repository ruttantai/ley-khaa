"""Did the script actually do the job? (spec §5.10, §6)

Pure: no I/O beyond reading what the run produced, no model call, no state
change. Every rule is recorded in `checks` for the manifest; only the first
failure in report order becomes the `reason` a human sees, and that reason is
always plain English — the traceback lives in the bundle, not in the question.
"""
from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path

from ..interpreter.spec import TaskSpec
from .formats import expected_suffixes
from .sandbox import SandboxResult
from .workspace import Workspace

# Suffixes we know how to count rows in. Anything else is not row-checked.
_TABULAR = {".csv", ".xlsx"}


@dataclass(frozen=True)
class Verdict:
    ok: bool
    reason: str
    checks: dict[str, bool]


def _format_matches(spec: TaskSpec, primary: Path | None) -> bool:
    if primary is None:
        return False
    suffixes = expected_suffixes(spec.output_format)
    # No opinion is a pass: rejecting a good deliverable because the request
    # described its format in words we did not anticipate is worse than not
    # checking at all.
    return not suffixes or primary.suffix.lower() in suffixes


def _has_rows(primary: Path | None) -> bool:
    if primary is None:
        return False
    suffix = primary.suffix.lower()
    if suffix not in _TABULAR:
        return True
    try:
        if suffix == ".csv":
            with primary.open(newline="", encoding="utf-8") as handle:
                return len(list(csv.reader(handle))) > 1
        from openpyxl import load_workbook

        book = load_workbook(primary)
        try:
            return book.active.max_row > 1
        finally:
            book.close()
    except Exception:
        # Unreadable is not "empty" but it is certainly not a good deliverable,
        # and the repair attempt is the right next move either way.
        return False


def validate(
    spec: TaskSpec,
    workspace: Workspace,
    result: SandboxResult,
    input_hashes: dict[str, str],
) -> Verdict:
    deliverables = workspace.deliverables()
    primary = deliverables[0] if deliverables else None
    # Links are reported, not followed. workspace.deliverables() already drops
    # them, so without this the bundle would simply look empty and the human
    # would be told "no output file" for a run that in fact tried to pass off a
    # file it never wrote.
    linked = workspace.linked_deliverables()

    checks = {
        "within_time_limit": not result.timed_out,
        "script_ran": result.exit_code == 0 and not result.timed_out,
        "deliverable_is_a_real_file": not linked,
        "deliverable_exists": primary is not None,
        "deliverable_not_empty": primary is not None and primary.stat().st_size > 0,
        "format_matches": _format_matches(spec, primary),
        "inputs_unmodified": workspace.input_hashes() == input_hashes,
        "has_rows": _has_rows(primary),
    }

    def fail(reason: str) -> Verdict:
        return Verdict(ok=False, reason=reason, checks=checks)

    if not checks["within_time_limit"]:
        return fail("The generated script ran too long and was stopped.")
    if not checks["script_ran"]:
        return fail("The generated script failed while running.")
    if not checks["deliverable_is_a_real_file"]:
        return fail(
            "The script left a link ("
            + ", ".join(path.name for path in linked)
            + ") in place of an output file instead of writing one."
        )
    if primary is None:
        # `checks["deliverable_exists"]` is exactly this test; asking the
        # question directly is what tells the rest of the function (and the
        # typechecker) that `primary` below is a real path.
        return fail("The script finished but produced no output file.")
    if not checks["deliverable_not_empty"]:
        return fail("The script produced an output file, but it is empty.")
    if not checks["format_matches"]:
        return fail(
            f"I was asked for {spec.output_format} but the script produced {primary.name}."
        )
    if not checks["inputs_unmodified"]:
        return fail(
            "The script changed its own input files, so the result cannot be reproduced."
        )
    if not checks["has_rows"]:
        return fail("The output file has a header but no rows in it.")

    return Verdict(
        ok=True,
        reason=f"Produced {primary.name} in {result.duration_ms} ms.",
        checks=checks,
    )
